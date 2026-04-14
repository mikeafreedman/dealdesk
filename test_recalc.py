"""
Quick test: run excel_builder with default assumptions, then read
the output with data_only=True to verify formulas were recalculated.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from models.models import (
    DealData, FinancialAssumptions, InvestmentStrategy, AssetType,
    PropertyAddress, RefiEvent, WaterfallTier, WaterfallType,
)
from excel_builder import populate_excel

# ── Build a minimal deal with realistic defaults ─────────────────────
refi_events = [
    RefiEvent(active=False, year=5, appraised_value=3200000, ltv=0.70,
              rate=0.06, amort_years=30, loan_term=10, orig_fee_pct=0.01,
              prepay_pct=0.01, closing_costs=25000),
    RefiEvent(active=False, year=8, appraised_value=3800000, ltv=0.65,
              rate=0.055, amort_years=30, loan_term=10, orig_fee_pct=0.01,
              prepay_pct=0.01, closing_costs=25000),
    RefiEvent(active=False, year=0, appraised_value=0, ltv=0.65,
              rate=0.055, amort_years=30, loan_term=10, orig_fee_pct=0.01,
              prepay_pct=0.0, closing_costs=0),
]

waterfall_tiers = [
    WaterfallTier(tier_number=1, hurdle_type="irr", hurdle_value=0.12, lp_share=0.70, gp_share=0.30),
    WaterfallTier(tier_number=2, hurdle_type="irr", hurdle_value=0.15, lp_share=0.60, gp_share=0.40),
    WaterfallTier(tier_number=3, hurdle_type="irr", hurdle_value=0.18, lp_share=0.30, gp_share=0.70),
    WaterfallTier(tier_number=4, hurdle_type="irr", hurdle_value=0.24, lp_share=0.20, gp_share=0.80),
]

assumptions = FinancialAssumptions(
    hold_period=10, num_units=20, gba_sf=15000, lot_sf=8000, year_built=1985,
    purchase_price=2500000, transfer_tax_rate=0.02139, closing_costs_fixed=75000,
    tenant_buyout=0, legal_closing=25000, title_insurance=8000, legal_bank=5000,
    appraisal=5000, environmental=6000, surveyor=3500, architect=0, structural=0,
    civil_eng=0, meps=0, legal_zoning=0, geotech=0,
    acq_fee_fixed=25000, mortgage_carry=0, mortgage_fees=17500, mezz_interest=0,
    working_capital=15000, marketing=5000, re_tax_carry=0, prop_ins_carry=0,
    dev_fee=0, dev_pref=0, permits=0, stormwater=0, demo=0, const_hard=0,
    const_reserve=0, gc_overhead=0, mezz_debt=0, tax_credit_equity=0, grants=0,
    ltv_pct=0.70, interest_rate=0.065, amort_years=30, loan_term=10,
    origination_fee_pct=0.01, io_period_months=0,
    refi_events=refi_events,
    vacancy_rate=0.075, annual_rent_growth=0.03, expense_growth_rate=0.03,
    loss_to_lease=0.03, cam_reimbursements=0, fee_income=6000,
    re_taxes=45000, insurance=18000, gas=12000, water_sewer=14000,
    electric=10000, license_inspections=2500, trash=8000,
    mgmt_fee_pct=0.06, salaries=24000, repairs=8000, exterminator=3600,
    cleaning=6000, turnover=5000, advertising=4000, landscape_snow=6000,
    admin_legal_acct=5000, office_phone=3000, miscellaneous=2000,
    cap_reserve_per_unit=400, commissions_yr1=0, renovations_yr1=0,
    exit_cap_rate=0.07, disposition_costs_pct=0.02,
    gp_equity_pct=0.10, waterfall_type=WaterfallType(1),
    pref_return=0.08, simple_lp_split=0.80,
    waterfall_tiers=waterfall_tiers,
    em_hurdle_t1=2.0, em_hurdle_t2=2.5, em_hurdle_t3=3.0,
    sens_rent_growth_low=0.0, sens_rent_growth_high=0.05, sens_rent_growth_step=0.01,
    sens_exit_cap_low=0.055, sens_exit_cap_high=0.085, sens_exit_cap_step=0.005,
)

deal = DealData(
    deal_id="test_recalc_001",
    asset_type=AssetType.MULTIFAMILY,
    investment_strategy=InvestmentStrategy.STABILIZED_HOLD,
    address=PropertyAddress(
        street="123 Test St", city="Philadelphia", state="PA",
        zip_code="19103", full_address="123 Test St, Philadelphia, PA 19103",
    ),
)
deal.assumptions = assumptions

# ── Run the builder (includes LibreOffice recalc) ────────────────────
print("Running populate_excel ...")
output_path = populate_excel(deal)
print(f"Output: {output_path}\n")

# ── Read back with data_only=True ────────────────────────────────────
wb = openpyxl.load_workbook(str(output_path), data_only=True)

# --- Assumptions tab: key formula cells ---
print("=" * 60)
print("ASSUMPTIONS TAB — key formula cells")
print("=" * 60)
ws = wb["Assumptions"]
formula_cells = {
    "C19": "Transfer Tax $",
    "C50": "Mortgage Fees / Origination",
    "C71": "Initial Loan Amount",
    "C76": "Origination Fee $",
    "C78": "Monthly Payment",
    "C79": "Annual Debt Service",
    "C82": "Senior Debt",
}
for cell_ref, label in formula_cells.items():
    val = ws[cell_ref].value
    print(f"  {cell_ref:5s}  {label:30s}  =  {val}")

# Also print a sampling of input cells to confirm they wrote correctly
print("\n  --- Sample input cells ---")
input_cells = {
    "C5": "Property Name",
    "C9": "Num Units",
    "C17": "Purchase Price",
    "C70": "LTV %",
    "C72": "Interest Rate",
    "C134": "Vacancy Rate",
}
for cell_ref, label in input_cells.items():
    val = ws[cell_ref].value
    print(f"  {cell_ref:5s}  {label:30s}  =  {val}")

# --- Pro Forma tab: first 10 rows ---
print("\n" + "=" * 60)
print("PRO FORMA TAB — first 10 rows")
print("=" * 60)
sheet_names = wb.sheetnames
print(f"Available sheets: {sheet_names}\n")

# Try common pro forma sheet names
pf_name = None
for candidate in ["Pro Forma", "ProForma", "Pro_Forma", "CF", "Cash Flow"]:
    if candidate in sheet_names:
        pf_name = candidate
        break
if pf_name is None:
    # Pick the second sheet if it exists
    if len(sheet_names) > 1:
        pf_name = sheet_names[1]
        print(f"(No 'Pro Forma' sheet found, using second sheet: '{pf_name}')")

if pf_name:
    pf = wb[pf_name]
    max_col = min(pf.max_column or 15, 15)  # cap at col O
    for row in pf.iter_rows(min_row=1, max_row=10, max_col=max_col, values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is None:
                vals.append("")
            elif isinstance(v, float):
                vals.append(f"{v:,.2f}")
            else:
                vals.append(str(v))
        print("  | ".join(f"{v:>14s}" for v in vals))
else:
    print("No secondary sheet found to inspect.")

wb.close()
print("\nDone.")
