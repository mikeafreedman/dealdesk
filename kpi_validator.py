"""
kpi_validator.py — PDF-vs-Excel KPI Diff Scaffold
===================================================
Diagnostic-only: reads 12 key KPIs from deal.financial_outputs (the
Python source of truth) AND from the on-disk Excel workbook (post-
write, post-recalculate), then logs any mismatch > $1 or 0.01%.

Purpose: lets us build a targeted cell-update list for a real
Excel/PDF sync pass in a follow-up. Not a fix in itself.

Call site: main.py, after excel_builder.populate_excel() and before
report_builder.generate_report().
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict

logger = logging.getLogger(__name__)

# KPI → (Python attribute, Excel sheet, cell address).  Cell addresses
# verified against hold_template_v3.xlsx by probing labels + adjacent
# value columns; the value column varies by sheet (B for Returns Summary,
# D for Cash Waterfall).
_KPI_MAP = [
    # Sources & Uses / equity
    ("total_project_cost",   "total_project_cost",     "Returns Summary",  "B7"),
    ("total_equity",         "total_equity_required",  "Cash Waterfall",   "D9"),
    ("gp_equity",            "gp_equity",              "Cash Waterfall",   "D7"),
    ("lp_equity",            "lp_equity",              "Cash Waterfall",   "D8"),
    ("initial_loan_amount",  "initial_loan_amount",    "Assumptions",      "C71"),
    # Operating year-1 metrics
    ("noi_yr1",              "noi_yr1",                "Pro Forma",        "B49"),
    ("going_in_cap_rate",    "going_in_cap_rate",      "Returns Summary",  "B13"),
    ("dscr_yr1",             "dscr_yr1",               "Returns Summary",  "B15"),
    ("cash_on_cash_yr1",     "cash_on_cash_yr1",       "Returns Summary",  "B14"),
    # Exit / return metrics
    ("project_irr",          "project_irr",            "Cash Waterfall",   "D30"),
    ("project_equity_multiple", "project_equity_multiple", "Cash Waterfall", "D31"),
    ("lp_irr",               "lp_irr",                 "Cash Waterfall",   "D41"),
    ("lp_equity_multiple",   "lp_equity_multiple",     "Cash Waterfall",   "D42"),
    ("gp_irr",               "gp_irr",                 "Cash Waterfall",   "D49"),
    ("gp_equity_multiple",   "gp_equity_multiple",     "Cash Waterfall",   "D50"),
]


def _close(py: Any, xl: Any) -> bool:
    """Values within $1 or 0.01% (relative) count as matching."""
    if py is None or xl is None:
        return py == xl
    try:
        pyf, xlf = float(py), float(xl)
    except (TypeError, ValueError):
        return str(py) == str(xl)
    if abs(pyf - xlf) <= 1.0:
        return True
    denom = max(abs(pyf), abs(xlf), 1.0)
    return abs(pyf - xlf) / denom <= 0.0001


def validate(deal, xlsx_path: str | Path) -> Dict[str, Dict[str, Any]]:
    """Compare Python-computed KPIs to the values currently in the xlsx.
    Logs every mismatch as a warning and returns the full diff as a dict
    so callers can surface it if they want.
    """
    from openpyxl import load_workbook

    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        logger.warning("KPI VALIDATOR: workbook not found at %s", xlsx_path)
        return {}

    try:
        # data_only=True so computed formula values come through.
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        logger.warning("KPI VALIDATOR: failed to open %s — %s", xlsx_path, exc)
        return {}

    fo = getattr(deal, "financial_outputs", None)
    if fo is None:
        logger.warning("KPI VALIDATOR: no financial_outputs on deal")
        return {}

    # Defensive cross-checks (#6, #8 from Excel audit): verify two
    # critical "live formula" cells match Python before we move on to the
    # KPI grid. If either of these drifts, every downstream KPI will
    # inherit the error, so we want a crisp standalone warning.
    def _xl(sheet: str, cell: str):
        try:
            return wb[sheet][cell].value
        except Exception:
            return None

    # (a) Rent Roll!E45 = Pro Forma!B6 = _gpr_yr1(deal). If this drifts,
    #     every year's GPR is wrong. B6 = ='Rent Roll'!E45 via formula.
    _gpr_xl = _xl("Rent Roll", "E45")
    _gpr_py = getattr(fo, "gross_potential_rent", None)
    if _gpr_xl is not None and _gpr_py is not None:
        if not _close(_gpr_py, _gpr_xl):
            logger.warning(
                "KPI PRECHECK GPR: Python=$%s vs Rent Roll!E45=$%s — "
                "Pro Forma year-1 GPR is wrong; all downstream income drifts",
                f"{_gpr_py:,.0f}", f"{_gpr_xl:,.0f}",
            )
        else:
            logger.info("KPI PRECHECK GPR: OK (Python=$%s == Rent Roll!E45)",
                        f"{_gpr_py:,.0f}")

    # (b) Assumptions!C89 (Total Uses) should now match Python.total_uses
    #     exactly because excel_builder overwrites C89 with fo.total_uses.
    _uses_xl = _xl("Assumptions", "C89")
    _uses_py = getattr(fo, "total_uses", None)
    if _uses_xl is not None and _uses_py is not None:
        if not _close(_uses_py, _uses_xl):
            logger.warning(
                "KPI PRECHECK Uses: Python=$%s vs Assumptions!C89=$%s — "
                "Sources & Uses balance is broken; equity cascade is off",
                f"{_uses_py:,.0f}", f"{_uses_xl:,.0f}",
            )
        else:
            logger.info("KPI PRECHECK Uses: OK (Python=$%s == Assumptions!C89)",
                        f"{_uses_py:,.0f}")

    diff: Dict[str, Dict[str, Any]] = {}
    for kpi_name, attr, sheet, cell in _KPI_MAP:
        py_val = getattr(fo, attr, None)
        if cell == "TODO" or sheet not in wb.sheetnames:
            xl_val = None
            note = f"cell-unknown (sheet={sheet})"
        else:
            try:
                xl_val = wb[sheet][cell].value
                note = f"{sheet}!{cell}"
            except Exception as exc:
                xl_val = None
                note = f"read-error: {exc}"
        matches = _close(py_val, xl_val)
        diff[kpi_name] = {"py": py_val, "xl": xl_val, "cell": note, "ok": matches}
        if not matches:
            logger.warning(
                "KPI DIFF %s: py=%s xl=%s (%s)",
                kpi_name, py_val, xl_val, note,
            )
        else:
            logger.info("KPI OK   %s: py=%s xl=%s", kpi_name, py_val, xl_val)
    wb.close()
    return diff
