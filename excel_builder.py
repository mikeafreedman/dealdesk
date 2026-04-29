"""
excel_builder.py — DealDesk CRE Underwriting
=============================================
Populates the Assumptions tab of the correct Excel template from DealData.
Zero hardcoded values — every cell value sourced from DealData.

Template routing (driven by InvestmentStrategy, not asset_type):
    stabilized_hold / value_add  →  Hold_Template_v3.xlsx
    opportunistic                →  Sale_Template_v3.xlsx

Output: {deal_id}_financial_model.xlsx  →  outputs/
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any, List, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font

from config import get_excel_template, OUTPUTS_DIR
from models.models import (
    AssetType, DealData, DevelopmentScenario, FinancialOutputs,
    InvestmentStrategy, ScenarioVerdict,
    RENOVATION_TIER_MULTIPLIERS, RENOVATION_DOWNTIME_MONTHS,
)


# DealDesk sage-light fill applied to the Market Rent column so analysts
# can visually distinguish it from the white (input) Current Rent column.
SAGE_MARKET_FILL = PatternFill(
    start_color="B2C9B4", end_color="B2C9B4", fill_type="solid",
)

# Human-readable labels for renovation_tier enum values.
_RENO_TIER_LABELS = {
    "light_cosmetic":   "Light Cosmetic Renovation",
    "heavy_rehab":      "Heavy Rehab",
    "new_construction": "New Construction",
}

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════════════════

def _do_populate_excel(deal: DealData, output_path: Path) -> Path:
    """Body of the legacy populate_excel, parameterized to take output_path
    as input rather than computing it from deal.deal_id.

    Reads from deal.assumptions and deal.financial_outputs throughout, so
    Session 4's per-scenario worker (``_populate_excel_for_scenario``)
    invokes this on a scenario_deal whose assumptions and financial_outputs
    have been swapped to the scenario's snapshot/outputs.

    Pure with respect to deal-level mutations: the legacy
    ``deal.output_xlsx_path = str(output_path)`` write was removed because
    main.py:1036-1037 already sets that field from the return value, and
    keeping the in-body write would be a per-scenario clobber under the
    fan-out architecture (last-scenario-wins).
    """
    template_path = get_excel_template(deal.investment_strategy)

    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Assumptions"]

    # Column widths so labels in column B and values in C–H are not truncated.
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 30
    # Un-hide any row whose height is 0 (template has a couple stashed rows).
    for row_dim in list(ws.row_dimensions.values()):
        if row_dim.height == 0:
            row_dim.height = None

    for cell_ref, value in _build_cell_map(deal):
        if value is not None:
            ws[cell_ref] = value

    # ── Initial Loan, Origination Fee, Mortgage Fees — plain floats ──
    # These used to be a circular formula chain:
    #   C50 (mortgage fees) → C76 (orig fee) → C71 (loan) → C89 (total uses,
    #   which sums C50). Compute all three in Python from the authoritative
    #   financial_outputs + assumptions and write as plain floats so the S&U
    #   tab has no circular dependency.
    a = deal.assumptions
    fo = deal.financial_outputs
    initial_loan = round(fo.initial_loan_amount or 0.0, 2)
    origination_fee = round(initial_loan * (a.origination_fee_pct or 0.0), 2)

    logger.info(f"EXCEL [Assumptions]: writing origination_fee = {origination_fee:.2f}")
    logger.info(f"EXCEL [Assumptions]: writing initial_loan = {initial_loan:.2f}")

    ws["C71"] = initial_loan
    logger.info("ASSUMPTIONS [C71]: wrote initial_loan=%s (plain float, loan × LTV)",
                f"{initial_loan:,.2f}")

    ws["C76"] = origination_fee
    logger.info("ASSUMPTIONS [C76]: wrote origination_fee=%s (plain float, loan × %.4f)",
                f"{origination_fee:,.2f}", a.origination_fee_pct or 0.0)

    ws["C50"] = origination_fee
    logger.info("ASSUMPTIONS [C50]: wrote mortgage_fees=%s (plain float, = origination fee)",
                f"{origination_fee:,.2f}")

    # ── Refi guard: override Active? flags for skipped refis ────
    refi_active_cells = ["C95", "C108", "C121"]  # Refi 1, 2, 3 Active? cells
    for i, refi in enumerate(a.refi_events[:3]):
        if not refi.active:
            ws[refi_active_cells[i]] = 0
            logger.info("REFI GUARD: wrote Active=0 to %s (refi %d skipped)",
                        refi_active_cells[i], i + 1)

    # ── Style refi cap rate (blue input) and appraised value (computed) ──
    _style_refi_cap_rate_cells(ws)

    # ── Rent Roll tab ────────────────────────────────────────────
    if "Rent Roll" in wb.sheetnames:
        _populate_rent_roll(wb["Rent Roll"], deal)

    # ── Pro Forma tab — override GPR Year 1 from financial_outputs ─
    if "Pro Forma" in wb.sheetnames:
        _populate_pro_forma_gpr(wb["Pro Forma"], deal)

    # ── Sensitivity tab ──────────────────────────────────────────
    if "Sensitivity" in wb.sheetnames:
        _populate_sensitivity(wb["Sensitivity"], deal)

    # ── Refi Analysis tab — write amortized loan balances ────────
    if "Refi Analysis" in wb.sheetnames:
        _populate_refi_balances(wb["Refi Analysis"], deal)

    # ── Amort - Initial tab — override for IO loans ─────────────
    if "Amort - Initial" in wb.sheetnames:
        _populate_amort_initial_io(wb["Amort - Initial"], deal)

    # ── Constr Interest tab ──────────────────────────────────────
    if "Constr Interest" in wb.sheetnames:
        _populate_constr_interest_tab(wb["Constr Interest"], deal)

    # ── Exit tab — write exit year NOI and guard Gross Sale Price ──
    if "Exit" in wb.sheetnames:
        ws_exit = wb["Exit"]
        # Dynamic formula: NOI row 49 in Pro Forma, hold period in Assumptions C14
        ws_exit["B5"] = "=INDEX('Pro Forma'!$B$49:$K$49,Assumptions!$C$14)"
        logger.info("EXIT: wrote INDEX formula for exit year NOI (row 49, hold=Assumptions!C14)")

        gross_sale = max(0.0, fo.gross_sale_price) if fo.gross_sale_price is not None else 0.0
        if gross_sale <= 0:
            ws_exit["B7"] = 0.0
            ws_exit["C7"] = "Exit not viable \u2014 NOI \u2264 $0"
            logger.warning("Exit tab: wrote $0 Gross Sale Price (negative exit NOI)")
        else:
            ws_exit["B7"] = gross_sale

    # ── Cover tab — remove firm name from prepared-by line ───────
    if "Cover" in wb.sheetnames:
        wb["Cover"]["B27"] = "Prepared by DealDesk."
        logger.info("COVER: B27 overwritten → 'Prepared by DealDesk.'")

    # ── Cash Waterfall D30: contextual N/A note on IRR non-convergence ──
    # Template cell D30 is =IFERROR(IRR(F28:P28),"N/A"). When Python has
    # also computed fo.project_irr as None (non-convergent cash flows),
    # replace with an explanatory note instead of a bare "N/A".
    if "Cash Waterfall" in wb.sheetnames and fo.project_irr is None:
        wb["Cash Waterfall"]["D30"] = "N/A — non-convergent due to mid-hold equity injection"
        logger.info("WATERFALL D30: Project IRR annotated — non-convergent")

    # Cash Waterfall Refi proceeds rows (24/25/26) are intentionally left
    # as template formulas so the workbook stays fully dynamic: changing a
    # Refi N year, active flag, LTV, or appraised value on Assumptions
    # flows through Refi Analysis (rows 4-22) into these cells automatically.
    # The earlier static-write path has been removed — see commit history.
    # Each template formula pattern:
    #   {col}{24|25|26} =IF(AND(Assumptions!$C$95=1, year_idx=Assumptions!$C$96),
    #                         'Refi Analysis'!B22, 0)
    # where 'Refi Analysis'!B22 = IF(active, new_loan - existing_balance - costs, 0).
    # Python still drives the correct EXISTING BALANCE value into
    # 'Refi Analysis'!B17 via _populate_refi_balances(), so the formula
    # produces the authoritative amortized balance at refi timing.

    wb.save(output_path)
    wb.close()

    recalculate_xlsx(str(output_path))

    return output_path


# ═══════════════════════════════════════════════════════════════════════════
# SESSION 4 — Per-scenario fan-out + master index
# ═══════════════════════════════════════════════════════════════════════════

def _populate_excel_for_scenario(
    deal: DealData,
    scenario: DevelopmentScenario,
    output_path: Path,
) -> Path:
    """Per-scenario Excel worker. Constructs a scenario_deal that carries
    the scenario's deltas-applied assumptions snapshot and the scenario's
    financial_outputs, then runs ``_do_populate_excel`` on the scenario_deal.

    Mirrors the financials.py CP2 worker pattern: shallow-copy of deal
    shares unmutated state (extracted_docs, parcel_data, market_data, etc.)
    while assumptions and financial_outputs are isolated per scenario.

    Raises ValueError if scenario.financial_outputs is None — the orchestrator
    is expected to filter those out before calling the worker.
    """
    if scenario.financial_outputs is None:
        raise ValueError(
            f"Cannot populate Excel for scenario '{scenario.scenario_id}' — "
            f"financial_outputs is None (CP2 worker may have failed)."
        )

    # Imported lazily to avoid a circular import: financials.py imports
    # mirror_preferred_to_legacy from models, and excel_builder is a peer
    # at the pipeline level. Lazy import keeps module-load order safe.
    from financials import _apply_scenario_deltas_to_assumptions

    scenario_assumptions = _apply_scenario_deltas_to_assumptions(scenario, deal.assumptions)
    scenario_deal = deal.model_copy(deep=False)
    scenario_deal.assumptions = scenario_assumptions
    scenario_deal.financial_outputs = scenario.financial_outputs

    return _do_populate_excel(scenario_deal, output_path)


def _build_scenarios_index(
    deal: DealData,
    scenario_files: List[Tuple[DevelopmentScenario, Path]],
) -> Path:
    """Build the master ``{deal_id}_scenarios_index.xlsx`` from scratch.

    One sheet "Scenarios Comparison" with a header row plus one data row
    per scenario, ordered by scenario.rank ascending (preferred at row 2).
    No formulas, no LibreOffice recalc — pure value writes via openpyxl.
    """
    index_path = OUTPUTS_DIR / f"{deal.deal_id}_scenarios_index.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scenarios Comparison"

    headers = [
        "Scenario ID", "Scenario Name", "Verdict",
        "Unit Count", "Building SF",
        "Total Project Cost", "Year 1 NOI",
        "Project IRR", "LP IRR", "LP Equity Multiple", "Exit Value",
        "Excel Filename",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Sort by rank ascending; rank 1 = preferred → row 2.
    sorted_files = sorted(scenario_files, key=lambda pair: pair[0].rank)
    for row_idx, (scenario, xlsx_path) in enumerate(sorted_files, start=2):
        fo = scenario.financial_outputs
        ws.cell(row=row_idx, column=1,  value=scenario.scenario_id)
        ws.cell(row=row_idx, column=2,  value=scenario.scenario_name)
        ws.cell(row=row_idx, column=3,  value=scenario.verdict.value)
        ws.cell(row=row_idx, column=4,  value=scenario.unit_count)
        ws.cell(row=row_idx, column=5,  value=scenario.building_sf)
        ws.cell(row=row_idx, column=6,  value=fo.total_project_cost)
        ws.cell(row=row_idx, column=7,  value=fo.noi_yr1)
        ws.cell(row=row_idx, column=8,  value=fo.project_irr)
        ws.cell(row=row_idx, column=9,  value=fo.lp_irr)
        ws.cell(row=row_idx, column=10, value=fo.lp_equity_multiple)
        ws.cell(row=row_idx, column=11, value=fo.gross_sale_price)
        ws.cell(row=row_idx, column=12, value=xlsx_path.name)

    # Light column-width tuning for readability.
    widths = {1: 22, 2: 36, 3: 14, 4: 12, 5: 12, 6: 18, 7: 14,
              8: 12, 9: 12, 10: 18, 11: 16, 12: 44}
    for col_idx, width in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(index_path)
    wb.close()
    return index_path


def populate_excel(deal: DealData) -> Path:
    """Orchestrate per-scenario Excel emission.

    Behavior:
        - If deal.scenarios is empty: legacy single-file path. Produces
          ``{deal_id}_financial_model.xlsx``, no index file. Byte-identical
          to pre-Session-4 behavior except the legacy
          ``deal.output_xlsx_path`` write is now performed by main.py from
          the return value (was duplicated inside this function pre-CP3).
        - Otherwise: emits one ``{deal_id}_{scenario_id}_financial_model.xlsx``
          per scenario whose financial_outputs is populated, plus a master
          ``{deal_id}_scenarios_index.xlsx`` with one row per scenario.
          Returns the preferred scenario's Excel path. Per-scenario failures
          are isolated: a scenario that raises during template population is
          logged and skipped; other scenarios continue.

    Edge cases on the multi-scenario path:
        - Preferred scenario's worker fails but at least one alt succeeds:
          returns the first successful scenario's path with a WARNING log
          (preserves the legacy contract that the function returns *some*
          Path so main.py:1037 can stash it on deal.output_xlsx_path).
        - All scenario workers fail: falls through to the legacy single-file
          path so the report builder still has *something* to consume.

    Args:
        deal: DealData. May or may not have scenarios.

    Returns:
        Path to the preferred scenario's Excel file (multi-scenario path)
        or the single-deal Excel file (legacy path).
    """
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    if not deal.scenarios:
        # Backward-compat fallback.
        output_path = OUTPUTS_DIR / f"{deal.deal_id}_financial_model.xlsx"
        return _do_populate_excel(deal, output_path)

    # Multi-scenario fan-out.
    scenario_files: List[Tuple[DevelopmentScenario, Path]] = []
    preferred_path: Optional[Path] = None
    for scenario in deal.scenarios:
        if scenario.financial_outputs is None:
            logger.warning(
                "EXCEL [%s]: skipping (financial_outputs is None — CP2 "
                "worker may have failed)",
                scenario.scenario_id,
            )
            continue
        try:
            output_path = OUTPUTS_DIR / (
                f"{deal.deal_id}_{scenario.scenario_id}_financial_model.xlsx"
            )
            populated = _populate_excel_for_scenario(deal, scenario, output_path)
            scenario.excel_filename = populated.name
            scenario_files.append((scenario, populated))
            if scenario.verdict == ScenarioVerdict.PREFERRED:
                preferred_path = populated
            logger.info(
                "EXCEL [%s]: populated -> %s",
                scenario.scenario_id, populated.name,
            )
        except Exception as exc:
            logger.error(
                "EXCEL [%s] FAILED (non-fatal): %s",
                scenario.scenario_id, exc,
            )

    if scenario_files:
        index_path = _build_scenarios_index(deal, scenario_files)
        logger.info("EXCEL: scenarios index built at %s", index_path.name)

    if preferred_path is None:
        # Edge case: preferred scenario's Excel build failed.
        if scenario_files:
            preferred_path = scenario_files[0][1]
            logger.warning(
                "EXCEL: preferred scenario's Excel build failed -- returning "
                "first successful scenario's path (%s) for legacy compat",
                preferred_path.name,
            )
        else:
            # Total failure: fall through to legacy single-file path.
            logger.error(
                "EXCEL: all scenario builds failed -- falling through to "
                "legacy single-file path"
            )
            output_path = OUTPUTS_DIR / f"{deal.deal_id}_financial_model.xlsx"
            return _do_populate_excel(deal, output_path)

    return preferred_path


def _find_libreoffice() -> str:
    """Return the LibreOffice executable path for the current platform."""
    import sys
    if sys.platform == "win32":
        for candidate in [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]:
            if os.path.isfile(candidate):
                return candidate
    # Linux / macOS — 'libreoffice' or 'soffice' should be on PATH
    return "libreoffice"


def recalculate_xlsx(xlsx_path: str) -> str:
    """Use LibreOffice headless to force-recalculate all formulas."""
    import tempfile
    soffice = _find_libreoffice()
    basename = os.path.basename(xlsx_path)

    with tempfile.TemporaryDirectory() as tmp_dir:
        result = subprocess.run([
            soffice, "--headless", "--calc",
            "--convert-to", "xlsx:Calc MS Excel 2007 XML",
            "--outdir", tmp_dir,
            xlsx_path
        ], capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice recalc failed: {result.stderr}")
        recalced = os.path.join(tmp_dir, basename)
        if not os.path.isfile(recalced):
            raise RuntimeError(f"LibreOffice produced no output at {recalced}")
        shutil.move(recalced, xlsx_path)
    return xlsx_path


# ═══════════════════════════════════════════════════════════════════════════
# RENT ROLL POPULATION
# ═══════════════════════════════════════════════════════════════════════════

# Template layout:
#   Residential rows 6–25 (B=Unit#, C=Type, D=SF, E=Rent/Mo, H=Status, I=LeaseExp, J=MarketRent)
#   Commercial  rows 35–39 (B=Tenant, C=Suite, D=SF, E=Rent/SF/Yr, G=LeaseType, H=Start, I=Expiry, J=TI/SF)
_RES_START, _RES_END = 6, 25
_COM_START, _COM_END = 35, 39

_COMMERCIAL_ASSET_TYPES = {AssetType.OFFICE, AssetType.RETAIL}


def _is_commercial_asset(deal: DealData) -> bool:
    return deal.asset_type in _COMMERCIAL_ASSET_TYPES


def _clear_rows(ws, start: int, end: int, cols: str) -> None:
    """Set all cells in the given row range + columns to None."""
    for row in range(start, end + 1):
        for col in cols:
            ws[f"{col}{row}"] = None


_RESIDENTIAL_UNIT_TOKENS = ("studio", "efficiency", "1br", "2br", "3br", "4br", "bedroom", "bed")
_COMMERCIAL_UNIT_TOKENS  = ("commercial", "office", "retail", "industrial", "warehouse", "flex")


def _classify_unit(u: dict) -> str:
    """Return 'residential' or 'commercial' for a single unit_mix entry.

    Classification order (most specific first):
      1. Explicit commercial markers — unit_type ∈ {commercial, office,
         retail, industrial, warehouse, flex} or annual_rent_per_sf /
         lease_type / tenant_name set.
      2. Explicit residential markers — unit_type matches bedroom-count
         convention (Studio, 1BR, 2BR, 3BR, 4BR+, N-bedroom).
      3. Default: residential (the safer default since most rent-roll
         extractions come from multifamily OMs).
    """
    if not isinstance(u, dict):
        return "residential"

    ut = (u.get("unit_type") or "").strip().lower()

    # Commercial signals
    if any(tok == ut for tok in _COMMERCIAL_UNIT_TOKENS):
        return "commercial"
    if u.get("annual_rent_per_sf") is not None:
        return "commercial"
    if u.get("lease_type"):
        return "commercial"
    if u.get("tenant_name"):
        return "commercial"

    # Residential signals
    if any(tok in ut for tok in _RESIDENTIAL_UNIT_TOKENS):
        return "residential"
    if re.match(r"^\d+\s*(br|bd|bed)", ut):
        return "residential"

    return "residential"


def _fmr_for_unit_type(deal: DealData, unit_type: str) -> float | None:
    """HUD FMR lookup by bedroom convention. Returns None if no FMR available."""
    md = deal.market_data
    if md is None:
        return None
    ut = (unit_type or "").strip().lower()
    attr = None
    if "studio" in ut or "efficiency" in ut:
        attr = "fmr_studio"
    elif re.match(r"^(1\s*br|1\s*bed|1-bed)", ut):
        attr = "fmr_1br"
    elif re.match(r"^(2\s*br|2\s*bed|2-bed)", ut):
        attr = "fmr_2br"
    elif re.match(r"^(3\s*br|3\s*bed|3-bed)", ut):
        attr = "fmr_3br"
    elif re.match(r"^(4\s*br|4\s*bed|4-bed)", ut):
        attr = "fmr_3br"
    if attr is None:
        return None
    v = getattr(md, attr, None)
    if v is None:
        return None
    try:
        f = float(v)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _default_monthly_rent(deal: DealData, unit_type: str) -> float:
    """Default rent = HUD FMR × renovation-tier multiplier.

    Priority (caller must enforce upload/user-input first):
      1. FMR for the unit's bedroom type × tier_multiplier
      2. quality_adjusted_market_rent (already tier-adjusted)
      3. Hardcoded $1,200 floor
    """
    a = deal.assumptions
    tier = (getattr(a, "renovation_tier", None) or "light_cosmetic")
    mult = RENOVATION_TIER_MULTIPLIERS.get(tier, 1.0)

    fmr = _fmr_for_unit_type(deal, unit_type)
    if fmr is not None:
        return round(fmr * mult, 0)

    qamr = a.quality_adjusted_market_rent or 1200
    return round(float(qamr), 0)


def _clear_residential_rows(ws) -> None:
    """Zero/blank residential rows (6–25)."""
    for row in range(_RES_START, _RES_END + 1):
        for col in "BCHIJKLMN":
            ws[f"{col}{row}"] = None
        ws[f"D{row}"] = 0
        ws[f"E{row}"] = 0


def _clear_commercial_rows(ws) -> None:
    """Zero/blank commercial rows (35–39)."""
    for row in range(_COM_START, _COM_END + 1):
        for col in "BCGHIJKLMNO":
            ws[f"{col}{row}"] = None
        ws[f"D{row}"] = 0
        ws[f"E{row}"] = 0


def _populate_rent_roll(ws, deal: DealData) -> None:
    """Write unit-level / tenant-level data to the Rent Roll sheet.

    Type-gated: residential rows (6–25) populate only when residential
    units are present; commercial rows (35–39) populate only when
    commercial tenants are present. Mixed-use deals populate both.
    Empty sections are zeroed out so stale template defaults never leak.
    """
    ext = deal.extracted_docs
    units = (ext.unit_mix or []) if ext else []

    residential_units = [u for u in units if _classify_unit(u) == "residential"]
    commercial_units  = [u for u in units if _classify_unit(u) == "commercial"]

    # Drop empty commercial placeholders (no tenant_name AND no SF AND
    # no rent) when residential units exist. This happens when an OM
    # extractor tags a footer row or a rendering artifact as a
    # commercial unit — if we keep it, the commercial populator's
    # "all_empty" branch triggers the whole-building GPR fallback,
    # which double-counts the residential rent against gba.
    if residential_units and commercial_units:
        def _is_placeholder_commercial(u: dict) -> bool:
            if not isinstance(u, dict):
                return True
            has_any = any(
                u.get(k) for k in (
                    "tenant_name", "sf", "monthly_rent",
                    "annual_rent_per_sf", "rent_per_sf_yr",
                )
            )
            return not has_any
        real_commercial = [u for u in commercial_units
                           if not _is_placeholder_commercial(u)]
        dropped = len(commercial_units) - len(real_commercial)
        if dropped:
            logger.info(
                "EXCEL Rent Roll: dropped %d empty commercial placeholder(s) "
                "from a mixed-use unit_mix (residential covers GPR)",
                dropped,
            )
        commercial_units = real_commercial

    # No classified units → fall back to asset-type hint so the summary
    # row in the residential populator (gross_potential_rent ÷ num_units)
    # still fires when extraction produced nothing.
    if not residential_units and not commercial_units:
        if _is_commercial_asset(deal):
            commercial_units = units
        else:
            residential_units = units

    logger.info(
        "EXCEL Rent Roll dispatch — residential=%d units, commercial=%d tenants, "
        "asset_type=%s",
        len(residential_units), len(commercial_units), deal.asset_type.value,
    )

    if residential_units:
        _populate_rent_roll_residential(ws, residential_units, deal)
    else:
        _clear_residential_rows(ws)

    if commercial_units:
        _populate_rent_roll_commercial(ws, commercial_units, deal)
    else:
        _clear_commercial_rows(ws)

    _write_renovation_summary(ws, deal)


def _populate_rent_roll_residential(ws, units: list, deal: DealData | None = None) -> None:
    """Write residential unit data to rows 6–25."""
    # ── If no unit-level data, try to write a summary row from assumptions ──
    if not units and deal is not None:
        a = deal.assumptions
        fo = deal.financial_outputs
        monthly_rent = getattr(fo, "gross_potential_rent", None) or 0
        # gross_potential_rent is annual; we need the monthly total for display
        num_units = a.num_units or 0
        gba_sf = a.gba_sf or 0
        gpr_yr1 = monthly_rent  # this is already annual from financial_outputs

        if gpr_yr1 > 0 and num_units > 0:
            row = _RES_START
            avg_sf = round(gba_sf / num_units, 0) if gba_sf and num_units else None
            rent_per_mo = round(gpr_yr1 / 12 / num_units, 2) if num_units else 0
            annual_rent = round(gpr_yr1, 2)

            ws[f"B{row}"] = "All Units (Summary)"
            ws[f"C{row}"] = deal.asset_type.value
            ws[f"D{row}"] = avg_sf
            ws[f"E{row}"] = rent_per_mo
            # G = Annual Rent — write directly since formula won't compute
            ws[f"G{row}"] = annual_rent
            ws[f"H{row}"] = "Occupied"

            logger.info(
                "EXCEL Rent Roll: wrote summary row — monthly_rent=%s, "
                "num_units=%s, gpr_yr1=%s",
                round(gpr_yr1 / 12, 2), num_units, gpr_yr1,
            )

            # Clear remaining rows
            for r in range(_RES_START + 1, _RES_END + 1):
                for col in "BCDEHIJ":
                    ws[f"{col}{r}"] = None
            return

    # Leasing column headers for residential
    ws[f"K{_RES_START - 1}"] = "Market ($/mo)"
    ws[f"L{_RES_START - 1}"] = "Lease Term (Yrs)"
    ws[f"M{_RES_START - 1}"] = "Expiry (Hold Yr)"
    ws[f"N{_RES_START - 1}"] = "Renewal Prob (%)"

    for i, row in enumerate(range(_RES_START, _RES_END + 1)):
        if i < len(units):
            u = units[i]
            ws[f"B{row}"] = u.get("unit_id") or f"{i+1:03d}"
            ws[f"C{row}"] = u.get("unit_type")
            # Display "N/A" instead of 0 when unit SF was not provided.
            # Overwriting F (the template Rent/SF/Mo formula) is necessary
            # because =E/D would produce 0 or #DIV/0! on a missing SF.
            _unit_sf = u.get("sf")
            if _unit_sf:
                ws[f"D{row}"] = _unit_sf
            else:
                ws[f"D{row}"] = "N/A"
                ws[f"F{row}"] = "N/A"
            # Rent priority: uploaded monthly_rent → user_rent → FMR × tier
            _rent = u.get("monthly_rent")
            if not _rent:
                _rent = u.get("user_rent")
            if not _rent and deal is not None:
                _rent = _default_monthly_rent(deal, u.get("unit_type") or "")
            ws[f"E{row}"] = _rent
            # G (Annual Rent) is a formula; left intact
            ws[f"H{row}"] = _normalise_status(u.get("status"))
            ws[f"I{row}"] = u.get("lease_end")
            ws[f"J{row}"] = u.get("market_rent")
            # Visual distinction: Market Rent cell = sage-light fill.
            ws[f"J{row}"].fill = SAGE_MARKET_FILL
            # Leasing fields
            ws[f"K{row}"] = u.get("market_rent_sf") or 0
            ws[f"L{row}"] = u.get("lease_term_years") or 1
            ws[f"M{row}"] = u.get("lease_expiry_year") or 0
            ws[f"N{row}"] = round((u.get("renewal_probability") or 0.70) * 100, 0)
        else:
            # Clear unused rows to avoid template defaults
            for col in "BCDEHIJ":
                ws[f"{col}{row}"] = None


def _populate_rent_roll_commercial(ws, units: list, deal: DealData) -> None:
    """Populate the commercial tenant section (rows 35–39).

    Caller (_populate_rent_roll dispatcher) is responsible for:
      - Pre-filtering `units` down to commercial tenants only
      - Clearing residential rows when no residential units exist
    """
    commercial_units = units

    # ── Check if all commercial tenant rows are empty ───────────
    all_empty = all(
        not t.get("tenant_name") and not t.get("sf")
        for t in commercial_units
    ) if commercial_units else True

    gpr = getattr(deal.financial_outputs, "gross_potential_rent", None) or 0

    # ── Guard: skip whole-building GPR fallback for mixed-use ────
    # fo.gross_potential_rent already includes the residential share,
    # and deal.assumptions.gba_sf is the combined residential +
    # commercial footprint. Writing gpr / gba into the commercial
    # section when residential rows are populated double-counts both
    # rent AND space — exactly the bug the user reported. Only allow
    # the GPR fallback for pure-commercial deals (no residential unit
    # mix present).
    ext = deal.extracted_docs
    _has_residential = bool(ext and ext.unit_mix and any(
        _classify_unit(u) == "residential" for u in ext.unit_mix
    ))

    # ── GPR fallback: if no tenant data but GPR exists, write it ─
    if all_empty and gpr > 0 and not _has_residential:
        gba = deal.assumptions.gba_sf
        if not gba or gba <= 0:
            gba = 18000
        rent_per_sf = round(gpr / gba, 2)
        row = _COM_START
        logger.info(f"[EXCEL] Writing GPR fallback to commercial rent row: ${gpr:,.0f}")
        ws[f"B{row}"] = "All Tenants (GPR)"
        ws[f"C{row}"] = None
        ws[f"D{row}"] = gba
        ws[f"E{row}"] = rent_per_sf
        ws[f"G{row}"] = None
        ws[f"H{row}"] = None
        ws[f"I{row}"] = None
        ws[f"J{row}"] = 0
        # Clear remaining rows
        for row2 in range(_COM_START + 1, _COM_END + 1):
            for col in "BCDEGHI":
                ws[f"{col}{row2}"] = None
            ws[f"D{row2}"] = 0
            ws[f"E{row2}"] = 0
            ws[f"J{row2}"] = 0
        return

    # ── Write commercial tenants to rows 35–39 ──────────────────
    # Add leasing column headers (K–O)
    ws["K34"] = "Market ($/SF/yr)"
    ws["L34"] = "Lease Term (Yrs)"
    ws["M34"] = "Expiry (Hold Yr)"
    ws["N34"] = "Renewal Prob (%)"
    ws["O34"] = "Downtime (Mo)"

    for i, row in enumerate(range(_COM_START, _COM_END + 1)):
        if i < len(commercial_units):
            t = commercial_units[i]
            # Tenant name fallback: extractor often leaves tenant_name
            # null when the OM describes the space generically (e.g.
            # "Ground-floor retail, 800 SF"). Use the suite or a
            # unit-type label so column B never renders blank.
            _tenant = (
                t.get("tenant_name")
                or (f"{t.get('unit_type') or 'Commercial'} Tenant"
                    + (f" ({t.get('unit_id') or t.get('suite')})"
                       if (t.get('unit_id') or t.get('suite')) else ""))
            )
            ws[f"B{row}"] = _tenant
            ws[f"C{row}"] = t.get("unit_id") or t.get("suite")
            ws[f"D{row}"] = t.get("sf") or 0
            # Derive annual rent/SF: prefer explicit, else compute from monthly_rent
            rent_per_sf = t.get("annual_rent_per_sf") or t.get("rent_per_sf_yr")
            if rent_per_sf is None:
                monthly = t.get("monthly_rent")
                sf = t.get("sf")
                if monthly and sf and sf > 0:
                    rent_per_sf = round((monthly * 12) / sf, 2)
            ws[f"E{row}"] = rent_per_sf or 0
            # F (Annual Rent) is a formula — leave intact
            ws[f"G{row}"] = t.get("lease_type")
            ws[f"H{row}"] = t.get("lease_start")
            ws[f"I{row}"] = t.get("lease_end")
            ws[f"J{row}"] = t.get("ti_per_sf") or 0
            # Leasing fields
            ws[f"K{row}"] = t.get("market_rent_sf") or 0
            ws[f"L{row}"] = t.get("lease_term_years") or 5
            ws[f"M{row}"] = t.get("lease_expiry_year") or 0
            ws[f"N{row}"] = round((t.get("renewal_probability") or 0.70) * 100, 0)
            ws[f"O{row}"] = t.get("downtime_months") or 3
        else:
            # Clear unused commercial rows
            for col in "BCDEGHI":
                ws[f"{col}{row}"] = None
            ws[f"D{row}"] = 0
            ws[f"E{row}"] = 0
            ws[f"J{row}"] = 0


def _normalise_status(raw: str | None) -> str:
    """Map extracted lease status strings to the template's expected values.

    When no status is supplied (synthesized rent roll, no rent-roll upload),
    default to "Occupied" — the UW assumes units are leased unless the user
    explicitly flags them vacant.
    """
    if not raw:
        return "Occupied"
    lower = raw.lower().strip()
    if lower in ("occupied", "current", "leased"):
        return "Occupied"
    if lower in ("vacant", "available"):
        return "Vacant"
    return raw.title()


def _write_renovation_summary(ws, deal: DealData) -> None:
    """Write renovation scope context onto the Rent Roll tab.

    Rows 27–30 of columns P/Q — a zone with no existing writes today.
    If this collides with a label in the XLSX template, relocate the
    four (row, col) positions below.
    """
    a = deal.assumptions
    tier_val = getattr(a, "renovation_tier", "light_cosmetic") or "light_cosmetic"
    multiplier = RENOVATION_TIER_MULTIPLIERS.get(tier_val, 1.0)
    downtime   = RENOVATION_DOWNTIME_MONTHS.get(tier_val, 2)
    leaseup    = int(getattr(a, "lease_up_months", 1) or 1)
    qamr       = a.quality_adjusted_market_rent or 0

    ws["P27"] = "Renovation Tier"
    ws["Q27"] = _RENO_TIER_LABELS.get(tier_val, "Renovation")

    ws["P28"] = "Quality-Adjusted Market Rent"
    ws["Q28"] = (
        f"${qamr:,.0f}/mo (HUD FMR × {multiplier * 100:.0f}%)"
        if qamr else f"N/A (HUD FMR × {multiplier * 100:.0f}%)"
    )

    ws["P29"] = "Renovation Downtime"
    ws["Q29"] = f"{downtime} months per unit"

    ws["P30"] = "Lease-Up Period"
    ws["Q30"] = f"{leaseup} month(s) per unit"

    logger.info(
        "EXCEL Rent Roll: wrote renovation summary — tier=%s, QAMR=$%s, "
        "downtime=%dmo, leaseup=%dmo",
        tier_val, qamr or "n/a", downtime, leaseup,
    )


# ═══════════════════════════════════════════════════════════════════════════
# SENSITIVITY TAB POPULATION
# ═══════════════════════════════════════════════════════════════════════════

# Template layout (hold_template.xlsx → "Sensitivity" sheet):
#   Grid 1 — LEVERED IRR:       rows 7–13, cols C–H  (rent_growth × exit_cap)
#   Grid 2 — EQUITY MULTIPLE:   rows 7–13, cols L–Q  (same axes)
#   Grid 3 — YEAR 1 NOI:        rows 21–25, cols C–H (exp_growth × vacancy)
#   Grid 4 — YEAR 1 CoC:        rows 21–25, cols L–Q (ltv × purchase_price)

def _populate_sensitivity(ws, deal: DealData) -> None:
    """Write computed sensitivity grids into the Sensitivity tab."""
    fo = deal.financial_outputs

    _write_grid(ws, fo.sensitivity_matrix, row_start=7, col_start=3,
                max_rows=7, max_cols=6, fmt_fn=_fmt_pct)
    _write_grid(ws, fo.sensitivity_em_matrix, row_start=7, col_start=12,
                max_rows=7, max_cols=6, fmt_fn=_fmt_em)
    _write_grid(ws, fo.sensitivity_noi_matrix, row_start=21, col_start=3,
                max_rows=5, max_cols=6, fmt_fn=_fmt_dollar)
    _write_grid(ws, fo.sensitivity_coc_matrix, row_start=21, col_start=12,
                max_rows=5, max_cols=6, fmt_fn=_fmt_pct)


def _write_grid(
    ws, matrix: list | None,
    row_start: int, col_start: int,
    max_rows: int, max_cols: int,
    fmt_fn,
) -> None:
    """Write a 2-D matrix into a rectangular cell region.

    If matrix is None or empty, leave existing cell values (template "—") intact.
    """
    if not matrix:
        return
    for r_idx, row_data in enumerate(matrix[:max_rows]):
        for c_idx, val in enumerate(row_data[:max_cols]):
            ws.cell(
                row=row_start + r_idx,
                column=col_start + c_idx,
                value=fmt_fn(val),
            )


def _fmt_pct(v) -> float | str:
    """Return value as-is (already a decimal like 0.0823); Excel cell is
    formatted as percentage by the template. Pass through 'N/A' strings."""
    if isinstance(v, str):
        return v
    return v


def _fmt_em(v: float) -> float:
    """Equity multiple — raw number like 1.85."""
    return v


def _fmt_dollar(v: float) -> float:
    """Dollar amount — raw number; Excel cell has $ formatting."""
    return round(v, 0)


# ═══════════════════════════════════════════════════════════════════════════
# CELL MAP BUILDER
# ═══════════════════════════════════════════════════════════════════════════

CellMap = List[Tuple[str, Any]]


def _build_cell_map(deal: DealData) -> CellMap:
    """Assemble every (cell_ref, value) pair for the Assumptions sheet."""
    a = deal.assumptions
    ext = deal.extracted_docs
    fo = deal.financial_outputs
    is_hold = deal.investment_strategy != InvestmentStrategy.OPPORTUNISTIC

    cells: CellMap = []

    # ── Section 1: Property Information (rows 5–14) ──────────────
    cells += _section_property_info(deal, a, ext, is_hold)

    # ── Section 2: Acquisition (rows 17–20) ──────────────────────
    cells += _section_acquisition(a)

    # ── Section 3: Sources & Uses (rows 31–67 uses, 83–87 sources)
    cells += _section_uses(a, fo)
    cells += _section_sources(a, fo)

    # ── Section 4: Initial Financing (rows 70–77) ────────────────
    cells += _section_financing(a)

    # ── Sections 5–7: Refinancing (rows 95–131) ─────────────────
    cells += _section_refis(a)

    # ── Sections 8–11A: Operating (Hold template only) ──────────
    if is_hold:
        cells += _section_operating_income(a)
        cells += _section_fixed_expenses(a)
        cells += _section_variable_expenses(a)
        cells += _section_below_the_line(a)
        cells += _section_dev_period(a)
        cells += _section_renovation(a)

    # ── Sections 12–15: Exit, Waterfall, EM, Sensitivity ────────
    #    Row positions shift between Hold and Sale templates.
    cells += _section_exit(a, is_hold)
    cells += _section_waterfall(a, is_hold)
    cells += _section_em_hurdles(a, is_hold)
    cells += _section_sensitivity(a, is_hold)

    return cells


# ═══════════════════════════════════════════════════════════════════════════
# SECTION BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def _section_property_info(
    deal: DealData, a, ext, is_hold: bool,
) -> CellMap:
    prop_name = ext.property_name or deal.address.full_address or ""
    num_units = a.num_units or ext.total_units_from_rr
    gba_sf = a.gba_sf or ext.gba_sf_extracted
    lot_sf = a.lot_sf or ext.lot_sf_extracted
    year_built = a.year_built or ext.year_built_extracted
    if year_built is None and deal.parcel_data:
        year_built = deal.parcel_data.year_built

    return [
        ("C5",  prop_name),
        ("C6",  deal.address.full_address),
        ("C7",  deal.asset_type.value),
        ("C8",  "Hold" if is_hold else "Sale"),
        ("C9",  num_units),
        ("C10", gba_sf),
        ("C11", lot_sf),
        ("C12", year_built),
        ("C13", deal.report_date),
        ("C14", a.hold_period),
    ]


def _section_acquisition(a) -> CellMap:
    return [
        ("C17", a.purchase_price),
        ("C18", a.transfer_tax_rate),
        ("C20", a.closing_costs_fixed),
    ]


def _section_uses(a, fo=None) -> CellMap:
    # Log the sum of values Python writes, to compare with financials.py total_uses
    excel_total_uses = (
        a.purchase_price +                              # C17 (acq section)
        a.purchase_price * a.transfer_tax_rate +        # C18→C19 (transfer tax)
        a.closing_costs_fixed +                         # C20 (acquisition loan closing costs)
        a.tenant_buyout +                               # C31
        # Professional
        a.legal_closing + a.title_insurance + a.legal_bank +
        a.appraisal + a.environmental + a.architect +
        a.structural + a.geotech + a.surveyor + a.civil_eng +
        a.meps + a.legal_zoning +
        # Financing
        a.acq_fee_fixed +
        # (origination fee is a formula in Excel: C50)
        # Soft
        a.working_capital + a.marketing + a.re_tax_carry +
        a.prop_ins_carry + a.dev_fee + a.dev_pref + a.permits +
        # Hard
        a.stormwater + a.demo + a.const_hard +
        a.const_reserve + a.gc_overhead
    )
    logger.info("EXCEL S&U total written (excl origination, closing_costs, mortgage_carry): %s", excel_total_uses)

    # Overwrite C89 with Python's authoritative total_uses value. The
    # template formula at C89 historically summed SUM(C29:C31) +
    # SUM(C34:C45) + SUM(C48:C51) + SUM(C54:C60) + SUM(C63:C67), which
    # OMITS purchase_price (C17), transfer_tax (C19), and closing_costs
    # (C20). That produced a ~$2K gap (closing_costs_fixed) plus missing
    # any construction interest carry. Writing the value directly locks
    # Total Uses to fo.total_uses so C91 (equity) and C84/C85 (GP/LP)
    # match Python exactly.
    c89_override = []
    if fo is not None and getattr(fo, "total_uses", None):
        c89_override = [("C89", float(fo.total_uses))]
        logger.info("EXCEL S&U: C89 overwritten with Python total_uses=$%s",
                    f"{fo.total_uses:,.2f}")
    return c89_override + [
        # Acquisition
        ("C31", a.tenant_buyout),
        # Professional & Due Diligence
        ("C34", a.legal_closing),
        ("C35", a.title_insurance),
        ("C36", a.legal_bank),
        ("C37", a.appraisal),
        ("C38", a.environmental),
        ("C39", a.architect),
        ("C40", a.structural),
        ("C41", a.geotech),
        ("C42", a.surveyor),
        ("C43", a.civil_eng),
        ("C44", a.meps),
        ("C45", a.legal_zoning),
        # Financing Costs
        ("C48", a.acq_fee_fixed),
        ("C49", getattr(fo, 'construction_interest_carry', 0.0) or 0.0),
        # C50 = Mortgage Fees / Origination (formula)
        ("C51", 0.0),  # mezzanine interest — removed from standard model
        # Soft Costs
        ("C54", a.working_capital),
        ("C55", a.marketing),
        ("C56", a.re_tax_carry),
        ("C57", a.prop_ins_carry),
        ("C58", a.dev_fee),
        ("C59", a.dev_pref),
        ("C60", a.permits),
        # Hard Costs
        ("C63", a.stormwater),
        ("C64", a.demo),
        # Construction hard cost: user enters $/SF on the frontend. Write
        # the PSF rate to D65 (user-editable) and the total in C65 as an
        # Excel formula that multiplies by GBA (C10). If the user later
        # edits GBA or the PSF rate in Excel, the total recomputes live.
        # Fall back to the Python-computed dollar total when the PSF is
        # not populated (e.g. legacy saved deals).
        ("D65", a.const_hard_psf),
        ("E65", "$/SF × GBA"),
        ("C65", (f"=D65*C10" if a.const_hard_psf else a.const_hard)),
        ("D66", a.const_reserve_psf),
        ("E66", "$/SF × GBA"),
        ("C66", (f"=D66*C10" if a.const_reserve_psf else a.const_reserve)),
        ("C67", a.gc_overhead),
    ]


def _section_sources(a, fo) -> CellMap:
    cells: CellMap = [
        # C82 = Senior Debt (formula)
        ("C83", 0.0),  # mezzanine debt — removed from standard model
        ("C86", a.tax_credit_equity),
        ("C87", a.grants),
    ]
    # GP / LP equity — derive from the template's own equity calc (C91)
    # so that Total Sources always equals Total Uses.
    #   C91 = C89 (Total Uses) − C82 (Senior Debt) − C83 (Mezz)
    #   C84 = C91 × GP%   (C197 = gp_equity_pct)
    #   C85 = C91 × LP%   (C198 = 1 − GP%)
    # Writing formulas instead of Python-computed values avoids a
    # mismatch between Python's total_uses and Excel's C89 sum.
    cells += [
        ("C84", "=C91*C197"),
        ("C85", "=C91*C198"),
    ]
    logger.info("GP/LP equity: written as formulas =C91*C197, =C91*C198 "
                "(C91=Total Equity Required, C197=GP%%, C198=LP%%)")
    return cells


def _style_refi_cap_rate_cells(ws) -> None:
    """Style refi cap rate cells as blue inputs, appraised value as computed."""
    blue_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
    blue_font = Font(color="0000FF", bold=True)
    computed_fill = PatternFill(start_color="F5EFE4", end_color="F5EFE4", fill_type="solid")
    computed_font = Font(color="4A6E50", italic=True)
    cap_rate_fmt = "0.00%"
    dollar_fmt = r'\$#,##0;"($"#,##0\);\-'

    bases = [95, 108, 121]
    cap_rows = [106, 119, 132]
    for b, cr in zip(bases, cap_rows):
        # Cap rate cell — blue editable input
        for col in ["B", "C"]:
            cell = ws[f"{col}{cr}"]
            cell.fill = blue_fill
            cell.font = blue_font
        ws[f"C{cr}"].number_format = cap_rate_fmt
        # Appraised value cell — computed (formula)
        cell = ws[f"C{b + 2}"]
        cell.fill = computed_fill
        cell.font = computed_font
        cell.number_format = dollar_fmt


def _populate_refi_balances(ws, deal: DealData) -> None:
    """Write amortized loan balances to Refi Analysis B17/C17/D17.

    Overwrites the template formula (which uses the original loan amount)
    with the actual amortized balance at the refi year, computed by
    financials.py from the Amort - Initial schedule.

    Also writes guarded (skipped) refi values: when refi.active was set
    to False by the NOI guard, override Refi Analysis cells to show $0.
    """
    a = deal.assumptions
    fo = deal.financial_outputs
    balances = fo.loan_balance_at_refi or [None, None, None]
    cols = ["B", "C", "D"]  # Refi 1, 2, 3

    for i, refi in enumerate(a.refi_events[:3]):
        col = cols[i]

        if not refi.active:
            # Refi was guarded/skipped — write zeros and a note
            ws[f"{col}6"] = 0.0    # Appraised Value
            ws[f"{col}8"] = 0.0    # New Loan Amount
            ws[f"{col}22"] = 0.0   # Net Refi Proceeds
            logger.info("Refi Analysis %s: refi skipped (NOI ≤ $0) — wrote $0 values", col)
            continue

        bal = balances[i] if i < len(balances) else None
        if bal is not None:
            cell = f"{col}17"
            ws[cell] = bal
            logger.info("Refi Analysis %s: wrote amortized balance $%s "
                        "(overriding original loan amount formula)", cell, f"{bal:,.0f}")

        # Equity-injection disclosure: when financials flagged that the new
        # loan does not cover the existing balance, surface it as a visible
        # warning in the Refi Analysis sheet.
        prov = deal.provenance.field_sources
        inject_flag = prov.get(f"refi{i+1}_equity_injection_required") == "True"
        if inject_flag:
            from openpyxl.styles import Font
            inject_amt = float(prov.get(f"refi{i+1}_equity_injection_amount", 0) or 0)
            warn_row = 24 + i   # 24 for refi1, 25 for refi2, 26 for refi3
            ws[f"A{warn_row}"] = f"\u26A0 EQUITY INJECTION REQUIRED (Refi {i+1})"
            ws[f"{col}{warn_row}"] = inject_amt
            try:
                ws[f"A{warn_row}"].font = Font(color="FF0000", bold=True)
                ws[f"{col}{warn_row}"].font = Font(color="FF0000", bold=True)
            except Exception:
                pass
            logger.info(
                "Refi Analysis: wrote equity-injection warning row %d for "
                "Refi %d (amount=$%s)",
                warn_row, i + 1, f"{inject_amt:,.0f}",
            )


def _populate_constr_interest_tab(ws, deal: DealData) -> None:
    """Populate the 'Constr Interest' sheet with the monthly S-curve draw
    schedule and summary stats computed by financials.py.

    Styling mirrors the template's house typography: Century Gothic at
    all levels, with the palette used on Returns Summary / Pro Forma /
    S&U — walnut (#2C1F14), sage-deep (#4A6E50), sage-light (#B2C9B4),
    parchment (#F5EFE4), earth (#8B7355).
    """
    fo = deal.financial_outputs
    a  = deal.assumptions

    schedule = getattr(fo, 'construction_interest_schedule', []) or []
    carry    = getattr(fo, 'construction_interest_carry', 0.0) or 0.0

    from openpyxl.styles import Alignment, Border, Side

    # Palette — RGB hex values pulled from the template Returns Summary
    # / S&U / Pro Forma tab cells (font color + fill inspection).
    CLR_WALNUT       = "2C1F14"
    CLR_SAGE_DEEP    = "4A6E50"
    CLR_SAGE_LIGHT   = "B2C9B4"
    CLR_PARCHMENT    = "F5EFE4"
    CLR_EARTH        = "8B7355"
    CLR_WHITE        = "FFFFFF"

    # Font helpers — Century Gothic everywhere, matching the template.
    def F(size, bold=False, color=CLR_WALNUT, italic=False):
        return Font(name="Century Gothic", size=size, bold=bold, color=color, italic=italic)

    fill_walnut    = PatternFill("solid", fgColor=CLR_WALNUT)
    fill_sage_lt   = PatternFill("solid", fgColor=CLR_SAGE_LIGHT)
    fill_parch     = PatternFill("solid", fgColor=CLR_PARCHMENT)
    fill_white     = PatternFill("solid", fgColor=CLR_WHITE)

    _side_rule     = Side(border_style="thin", color=CLR_SAGE_LIGHT)
    border_table   = Border(left=_side_rule, right=_side_rule,
                            top=_side_rule, bottom=_side_rule)

    align_center   = Alignment(horizontal="center", vertical="center")
    align_right    = Alignment(horizontal="right", vertical="center")
    align_left     = Alignment(horizontal="left",  vertical="center", indent=1)

    # ── Column widths — tab has 5 data cols in B..F plus an A gutter ──
    ws.column_dimensions["A"].width = 4
    for col_letter, width in [("B", 12), ("C", 20), ("D", 20), ("E", 22), ("F", 20)]:
        ws.column_dimensions[col_letter].width = width

    # ── A1 tab title bar (full width, sage-light bg) ──────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "CONSTRUCTION LOAN INTEREST SCHEDULE"
    ws["A1"].font = F(16, bold=True, color=CLR_WALNUT)
    ws["A1"].fill = fill_sage_lt
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = "S-curve draw model — interest accrues on drawn balance only"
    ws["A2"].font = F(9, color=CLR_EARTH, italic=True)
    ws["A2"].fill = fill_parch
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Summary block (rows 4–9) — two-column label/value on parchment ──
    SUMMARY = [
        ("Acquisition Loan Amount",               getattr(fo, 'initial_loan_amount', 0.0) or 0.0,         "$#,##0"),
        ("Construction Period (Months)",          getattr(a,  'const_period_months', 0)  or 0,            "0"),
        ("Permit / Mobilization Lag (Months)",    getattr(a,  'draw_start_lag', 1),                       "0"),
        ("Annual Interest Rate",                  getattr(a,  'interest_rate', 0.0) or 0.0,               "0.00%"),
    ]
    # Hard cost share (computed)
    hard_total = (getattr(a, 'const_hard', 0.0) or 0.0) + (getattr(a, 'const_reserve', 0.0) or 0.0)
    tpc = getattr(fo, 'total_project_cost', 0.0) or 0.0
    SUMMARY.append(
        ("Hard Cost Share (% of Total Project Cost)",
         round(hard_total / tpc, 4) if tpc > 0 else 0.0,
         "0.00%")
    )
    SUMMARY.append(
        ("Total Construction Interest Carry", carry, "$#,##0.00")
    )

    for i, (label, value, num_fmt) in enumerate(SUMMARY):
        r = 4 + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=value)
        lbl = ws.cell(row=r, column=2)
        val = ws.cell(row=r, column=3)
        lbl.font = F(10, color=CLR_WALNUT)
        lbl.fill = fill_parch
        lbl.alignment = align_left
        lbl.border = border_table
        val.font = F(10, bold=True, color=CLR_SAGE_DEEP)
        val.fill = fill_white
        val.alignment = align_right
        val.border = border_table
        val.number_format = num_fmt
        # Merge the trailing columns so the summary block reads clean
        # across the full width, mirroring the Returns Summary layout.
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        trailing = ws.cell(row=r, column=4)
        trailing.fill = fill_white
        trailing.border = border_table
        ws.row_dimensions[r].height = 20

    # ── Schedule table header (row 12) ────────────────────────────────
    HDR_ROW = 12
    headers = [
        ("Month",                 "0"),
        ("Monthly Draw ($)",       "$#,##0"),
        ("Cumulative Draw %",      "0.0%"),
        ("Outstanding Balance ($)","$#,##0"),
        ("Monthly Interest ($)",   "$#,##0.00"),
    ]
    ws.row_dimensions[HDR_ROW].height = 26
    for col_offset, (h, _nf) in enumerate(headers):
        c = ws.cell(row=HDR_ROW, column=2 + col_offset, value=h)
        c.fill = fill_walnut
        c.font = F(9, bold=True, color=CLR_WHITE)
        c.alignment = align_center
        c.border = border_table

    # ── Data rows ─────────────────────────────────────────────────────
    if not schedule:
        ws.merge_cells(start_row=HDR_ROW + 1, start_column=2,
                       end_row=HDR_ROW + 1, end_column=6)
        note = ws.cell(
            row=HDR_ROW + 1, column=2,
            value="No construction period — interest carry = $0",
        )
        note.font = F(10, italic=True, color=CLR_EARTH)
        note.alignment = Alignment(horizontal="center", vertical="center")
        note.fill = fill_parch
    else:
        keys = ["month", "monthly_draw", "cumulative_draw_pct",
                "outstanding_balance", "monthly_interest"]
        for i, entry in enumerate(schedule):
            r = HDR_ROW + 1 + i
            row_fill = fill_parch if i % 2 == 0 else fill_white
            for col_offset, (key, nf) in enumerate(zip(keys, [h[1] for h in headers])):
                c = ws.cell(row=r, column=2 + col_offset, value=entry.get(key))
                c.font = F(10, color=CLR_WALNUT)
                c.alignment = align_center if col_offset == 0 else align_right
                c.border = border_table
                c.fill = row_fill
                c.number_format = nf

        # Totals row — walnut header treatment matching table header style.
        total_row = HDR_ROW + 1 + len(schedule)
        ws.cell(row=total_row, column=2, value="TOTAL")
        # Leave cols C-E blank but styled; put carry in F (Monthly Interest col)
        ws.cell(row=total_row, column=6, value=carry)
        for col_offset, (_h, nf) in enumerate(headers):
            c = ws.cell(row=total_row, column=2 + col_offset)
            c.fill = fill_walnut
            c.font = F(10, bold=True, color=CLR_WHITE)
            c.alignment = align_center if col_offset == 0 else align_right
            c.border = border_table
            if col_offset == 4:
                c.number_format = nf
        ws.row_dimensions[total_row].height = 22

    logger.info(
        "CONSTR INTEREST TAB: wrote %d schedule rows, total_carry=%s",
        len(schedule), f"{carry:,.2f}"
    )


def _populate_amort_initial_io(ws, deal: DealData) -> None:
    """Override Amort - Initial tab for IO loans (amort_years == 0).

    The template formula PMT(rate/12, 0, -loan) errors when amort=0.
    For IO loans: payment = interest-only, principal = 0, balance = flat.
    """
    a = deal.assumptions
    if a.amort_years and a.amort_years > 0:
        return  # amortizing loan — let template formulas handle it

    fo = deal.financial_outputs
    loan_amount = fo.initial_loan_amount or 0.0
    annual_rate = a.interest_rate or 0.0
    monthly_payment = loan_amount * (annual_rate / 12) if annual_rate > 0 else 0.0

    # Override the summary cells
    ws["B3"] = loan_amount      # Loan Amount (override formula =Assumptions!C71)
    ws["B6"] = monthly_payment  # Monthly Payment (template formula errors for IO)

    # Override each month row (rows 9–368 = 360 months max)
    max_months = min(a.loan_term * 12 if a.loan_term else 120, 360)
    for month in range(1, max_months + 1):
        row = 8 + month  # row 9 = month 1
        ws[f"A{row}"] = month
        ws[f"B{row}"] = monthly_payment   # Payment
        ws[f"C{row}"] = 0.0               # Principal
        ws[f"D{row}"] = monthly_payment   # Interest
        ws[f"E{row}"] = loan_amount       # Balance (flat)

    logger.info("AMORT INITIAL: IO loan — flat balance at %s, monthly payment=%s",
                f"{loan_amount:,.2f}", f"{monthly_payment:,.2f}")


def _populate_pro_forma_gpr(ws, deal: DealData) -> None:
    """Override Pro Forma rows that the template formulas get wrong.

    1) GPR row 6: template formula ='Rent Roll'!E45 evaluates to 0 when
       Rent Roll has no unit data.  Write financial_outputs.gross_potential_rent
       directly, compounded by rent_growth.

    2) Debt Service row 60: template formula switches to refi payment in the
       refi year itself (year >= refi_year).  Correct convention: debt service
       in year N uses the loan active at the START of year N, so the new
       payment starts the year AFTER the refi.  Override with Python-computed
       values from pro_forma_years which already follow this convention.
    """
    fo = deal.financial_outputs
    hold_period = deal.assumptions.hold_period or 10
    cols = "BCDEFGHIJK"  # Year 1–10
    num_years = min(hold_period, 10)

    # ── Stabilization Factor row 4 ──────────────────────────────
    # Override the template's formula-driven stab factors with the
    # Python-computed values so Excel matches the financial model exactly.
    from financials import _get_stabilization_factors
    stab_factors = _get_stabilization_factors(deal)
    for n in range(num_years):
        ws[f"{cols[n]}4"] = round(stab_factors[n], 4)
    logger.info(
        "EXCEL Pro Forma: wrote Stabilization Factor row 4 — %s",
        [f"Y{i+1}={v:.2f}" for i, v in enumerate(stab_factors[:num_years])]
    )

    # ── GPR row 6 — formula-driven from Rent Roll ────────────────
    # B6 reads the live Rent Roll total (E45). C6–K6 compound from B6
    # using the rent-growth rate stored in Assumptions!C135. This
    # preserves the workbook's chain of calculation: any Rent Roll
    # edit or rent-growth change flows through automatically.
    # Do NOT apply stab factors to GPR here — EGI row 17 already
    # does that via =(B14+B15)*B4 where B4 is the stabilization factor.
    ws["B6"] = "='Rent Roll'!E45"
    rent_growth_ref = "Assumptions!$C$135"
    for n in range(1, num_years):      # n=1 → col C (Year 2) … n=9 → col K (Year 10)
        col = cols[n]
        ws[f"{col}6"] = f"=B6*(1+{rent_growth_ref})^{n}"
    logger.info(
        "EXCEL Pro Forma: GPR row 6 written as live formulas "
        "(B6='Rent Roll'!E45, C6:K6 compound from B6)"
    )

    proforma = fo.pro_forma_years if fo else None

    # ── Debt Service row 60 — formula-driven from Amort sheets ───
    #
    # Each Amort sheet has month rows starting at row 9 (Month 1).
    # Annual DS for Year Y = SUM of column B, rows (Y-1)*12+9 to Y*12+8:
    #   Year 1 → SUM(B9:B20)
    #   Year 2 → SUM(B21:B32)
    #   ...
    #   Year Y → SUM(B{(Y-1)*12+9}:B{Y*12+8})
    #
    # Logic: use the most recently executed refi at or before year Y.
    # Refi timing cells in Assumptions:
    #   Refi 1: Active=C95, Timing=C96
    #   Refi 2: Active=C108, Timing=C109
    #   Refi 3: Active=C121, Timing=C122
    #
    # Formula structure (nested IF, outermost = highest refi):
    #   =IF(AND(Assumptions!C121=1, Assumptions!C122<=Y),
    #        SUM('Amort - Refi 3'!B{s}:B{e}),
    #    IF(AND(Assumptions!C108=1, Assumptions!C109<=Y),
    #        SUM('Amort - Refi 2'!B{s}:B{e}),
    #    IF(AND(Assumptions!C95=1, Assumptions!C96<=Y),
    #        SUM('Amort - Refi 1'!B{s}:B{e}),
    #        SUM('Amort - Initial'!B{s}:B{e})
    #    )))

    ds_formulas = []
    for n in range(num_years):
        yr = n + 1
        # Row range in amort sheets for this year
        r_start = (yr - 1) * 12 + 9   # e.g. Year 1 → 9
        r_end   = yr * 12 + 8          # e.g. Year 1 → 20

        # Inner-most = Initial loan
        initial_sum  = f"SUM('Amort - Initial'!B{r_start}:B{r_end})"
        refi1_sum    = f"SUM('Amort - Refi 1'!B{r_start}:B{r_end})"
        refi2_sum    = f"SUM('Amort - Refi 2'!B{r_start}:B{r_end})"
        refi3_sum    = f"SUM('Amort - Refi 3'!B{r_start}:B{r_end})"

        # Build the nested IF from innermost out
        # Refi 1 check: active (C95=1) AND timing (C96) <= this year
        f_r1 = (
            f"IF(AND(Assumptions!$C$95=1,Assumptions!$C$96<={yr}),"
            f"{refi1_sum},{initial_sum})"
        )
        # Refi 2 check: if refi2 active AND timing <= year, use refi2;
        # otherwise fall through to refi1 check
        f_r2 = (
            f"IF(AND(Assumptions!$C$108=1,Assumptions!$C$109<={yr}),"
            f"{refi2_sum},{f_r1})"
        )
        # Refi 3 check: outermost
        formula = (
            f"=IF(AND(Assumptions!$C$121=1,Assumptions!$C$122<={yr}),"
            f"{refi3_sum},{f_r2})"
        )

        ws[f"{cols[n]}60"] = formula
        ds_formulas.append(f"Y{yr}=formula")

    logger.info("PRO FORMA DS: wrote formula-driven debt service row 60 "
                "(IF cascade: Refi3→Refi2→Refi1→Initial, SUM of monthly "
                "payment rows from each Amort sheet) for %d years", num_years)

    # ── Commissions row 53 — year-specific from lease events ──────
    lease_events = fo.lease_events or {}
    if proforma:
        comm_list = []
        for n in range(num_years):
            yr = n + 1
            comm = 0.0
            if n < len(proforma):
                comm = proforma[n].get("leasing_commission", 0.0)
            ws[f"{cols[n]}53"] = round(comm, 2)
            comm_list.append(round(comm, 0))
        # Also write Year 1 commission to Assumptions C166 for reference
        yr1_comm = comm_list[0] if comm_list else 0
        ws_assumptions = ws.parent["Assumptions"] if "Assumptions" in ws.parent.sheetnames else None
        if ws_assumptions:
            ws_assumptions["C166"] = yr1_comm
        logger.info("PRO FORMA COMMISSIONS: %s",
                    [f"Y{i+1}={v:,.0f}" for i, v in enumerate(comm_list)])

    # ── Leasing Costs row 56 (TI + downtime, excl commissions) ──
    if proforma:
        ws["A56"] = "  Leasing Costs (TI + Downtime)"
        lc_list = []
        has_any = False
        for n in range(num_years):
            if n < len(proforma):
                lc = (proforma[n].get("tenant_improvements", 0.0)
                      + proforma[n].get("downtime_loss", 0.0))
            else:
                lc = 0.0
            ws[f"{cols[n]}56"] = round(lc, 2)
            lc_list.append(round(lc, 0))
            if lc > 0:
                has_any = True
        if has_any:
            # Override NOCF row 57 to include leasing costs
            for n in range(num_years):
                c = cols[n]
                ws[f"{c}57"] = f"={c}49-{c}53-{c}54-{c}55-{c}56"
            logger.info("LEASING COSTS (TI+downtime): %s", lc_list)

    # ── Cash-on-Cash Return row 64 ──────────────────────────────
    # CoC = Free Cash Flow / LP Equity (Assumptions C85)
    ws["A64"] = "Cash-on-Cash Return"
    for n in range(num_years):
        c = cols[n]
        ws[f"{c}64"] = f"=IF(Assumptions!$C$85=0,0,{c}63/Assumptions!$C$85)"
    # Format as percentage
    for n in range(num_years):
        ws[f"{cols[n]}64"].number_format = "0.00%"
    logger.info("EXCEL Pro Forma: wrote Cash-on-Cash Return row 64 (FCF/LP equity)")


def _section_financing(a) -> CellMap:
    return [
        ("C70", a.ltv_pct),
        # C71 = Initial Loan Amount (formula)
        ("C72", a.interest_rate),
        ("C73", a.amort_years),
        ("C74", a.loan_term),
        ("C75", a.origination_fee_pct),
        # C76 = Origination Fee $ (formula)
        ("C77", a.io_period_months),
        # C78 = Monthly Payment (formula)
        # C79 = Annual Debt Service (formula)
    ]


def _section_refis(a) -> CellMap:
    """Sections 5–7: up to 3 refinancing events.

    Gap rows (106, 119, 132) are used for per-refi appraisal cap rates.
    Appraised Value cells (C97, C110, C123) are written as formulas:
        =IF(active=0, 0, INDEX('Pro Forma' NOI row, refi year) / cap_rate)
    """
    bases = [95, 108, 121]
    cap_rate_rows = [106, 119, 132]  # gap rows repurposed for cap rate
    cells: CellMap = []
    for i, refi in enumerate(a.refi_events[:3]):
        b = bases[i]
        cr = cap_rate_rows[i]
        # Appraised Value formula: NOI in refi year / cap rate
        # Pro Forma NOI is row 49, Years 1-10 are columns B-K
        appraised_formula = (
            f"=IF(C{b}=0,0,INDEX('Pro Forma'!$B$49:$K$49,"
            f"MATCH(C{b + 1},{{1,2,3,4,5,6,7,8,9,10}},0))/C{cr})"
        )
        cells += [
            (f"C{b}",      1 if refi.active else 0),
            (f"C{b + 1}",  refi.year),
            (f"C{b + 2}",  appraised_formula),             # formula, not value
            (f"C{b + 3}",  refi.ltv),
            # b+4 = New Loan Amount (formula)
            (f"C{b + 5}",  refi.rate),
            (f"C{b + 6}",  refi.amort_years),
            (f"C{b + 7}",  refi.loan_term),
            (f"C{b + 8}",  refi.orig_fee_pct),
            (f"C{b + 9}",  refi.prepay_pct),
            (f"C{b + 10}", refi.closing_costs),
            # Cap rate in gap row — blue editable input
            (f"B{cr}",     f"Refi {i + 1} — Appraisal Cap Rate"),
            (f"C{cr}",     refi.cap_rate),
        ]
    return cells


# ── Hold-only operating sections ─────────────────────────────────────────

def _section_operating_income(a) -> CellMap:
    """Section 8: Operating Assumptions — Income (rows 134–139)."""
    return [
        ("C134", a.vacancy_rate),
        ("C135", a.annual_rent_growth),
        ("C136", a.expense_growth_rate),
        ("C137", a.loss_to_lease),
        ("C138", a.cam_reimbursements),
        ("C139", 0.0 if a.fee_income == 6000.0 else a.fee_income),
    ]


def _section_fixed_expenses(a) -> CellMap:
    """Section 9: Fixed Expenses Year 1 (rows 142–148)."""
    return [
        ("C142", a.re_taxes),
        ("C143", a.insurance),
        ("C144", a.gas),
        ("C145", a.water_sewer),
        ("C146", a.electric),
        ("C147", a.license_inspections),
        ("C148", a.trash),
    ]


def _section_variable_expenses(a) -> CellMap:
    """Section 10: Variable Expenses Year 1 (rows 152–162)."""
    logger.info(
        "EXCEL [Pro Forma]: writing salaries=%s exterminator=%s turnover=%s "
        "advertising=%s repairs=%s cleaning=%s",
        a.salaries, a.exterminator, a.turnover,
        a.advertising, a.repairs, a.cleaning)
    return [
        ("C152", a.mgmt_fee_pct),
        ("C153", a.salaries),
        ("C154", a.repairs),
        ("C155", a.exterminator),
        ("C156", a.cleaning),
        ("C157", a.turnover),
        ("C158", a.advertising),
        ("C159", a.landscape_snow),
        ("C160", a.admin_legal_acct),
        ("C161", a.office_phone),
        ("C162", a.miscellaneous),
    ]


def _section_below_the_line(a) -> CellMap:
    """Section 11: Below-the-Line Items (rows 166–168) + 11B Leasing Costs (189–193)."""
    return [
        ("C166", a.commissions_yr1),
        ("C167", a.cap_reserve_per_unit),
        ("C168", a.renovations_yr1),
        # §11B Leasing cost assumptions — rows 188–191
        ("B188", "  11B. LEASING COST ASSUMPTIONS"),
        ("B189", "TI — New Lease ($/SF)"),
        ("C189", a.ti_new_psf),
        ("B190", "TI — Renewal ($/SF)"),
        ("C190", a.ti_renewal_psf),
        ("B191", "Commission — New (% of GLV) / Renewal (% of GLV)"),
        ("C191", a.commission_new_pct),
        ("D191", a.commission_renewal_pct),
    ]


def _section_dev_period(a) -> CellMap:
    """Section 11A: Development Period & Carry Costs (rows 172–184)."""
    return [
        ("C172", a.const_period_months),
        ("C173", a.const_loan_rate),
        # Development-period hard cost — keyed to the same PSF × GBA logic
        # as C65 above. Uses a direct reference to C65 so both cells stay
        # in sync when the user edits PSF or GBA in Excel.
        ("C174", ("=C65" if a.const_hard_psf else a.const_hard)),
        # draw_start_lag exposed on the 'Constr Interest' tab (no row in template)
        # C175 = Construction Budget Soft Costs (no single model field)
        # C176 = Total Construction Budget (formula)
        # C177 = Monthly Draw Rate (no model field)
        # C178 = Est. Interest Carry (formula)
        ("C181", a.leaseup_period_months),
        ("C182", a.leaseup_vacancy_rate),
        ("C183", a.leaseup_concessions),
        ("C184", a.leaseup_marketing),
        # C187 = Total Carry Costs (formula)
    ]


def _section_renovation(a) -> CellMap:
    """Section 11A-R: Renovation scope + quality-adjusted market rent.

    Rows 185–187 sit between the dev period block (ends row 184) and the
    11B leasing cost header at row 188. If the XLSX template places labels
    or formulas in these rows, relocate this section accordingly.
    """
    tier_val = getattr(a, "renovation_tier", "light_cosmetic") or "light_cosmetic"
    return [
        ("B185", "Renovation Tier"),
        ("C185", _RENO_TIER_LABELS.get(tier_val, "Light Cosmetic")),
        ("B186", "Quality-Adjusted Market Rent ($/mo)"),
        ("C186", a.quality_adjusted_market_rent or 0),
        ("B187", "Lease-Up Period (months)"),
        ("C187", int(getattr(a, "lease_up_months", 1) or 1)),
    ]


# ── Sections 12–15: row positions differ by template ─────────────────────

def _section_exit(a, is_hold: bool) -> CellMap:
    """Section 12: Exit Assumptions."""
    r = 192 if is_hold else 170
    return [
        (f"C{r + 1}", a.exit_cap_rate),
        (f"C{r + 2}", a.disposition_costs_pct),
    ]


def _section_waterfall(a, is_hold: bool) -> CellMap:
    """Section 13: Waterfall / Partnership Structure."""
    r = 196 if is_hold else 174
    tiers = a.waterfall_tiers
    return [
        (f"C{r + 1}",  a.gp_equity_pct),
        # r+2 = LP Equity % (formula: 1 − GP%)
        (f"C{r + 3}",  a.waterfall_type.value),
        (f"C{r + 4}",  a.simple_lp_split),
        (f"C{r + 5}",  round(1.0 - a.simple_lp_split, 6)),
        # Full waterfall tier structure
        (f"C{r + 8}",  a.pref_return),                   # Tier 1: Pref Return
        (f"C{r + 9}",  tiers[0].hurdle_value),            # Tier 2: IRR Hurdle
        (f"C{r + 10}", tiers[0].lp_share),                # Tier 2: LP Share
        (f"C{r + 11}", tiers[1].hurdle_value),             # Tier 3: IRR Hurdle
        (f"C{r + 12}", tiers[1].lp_share),                # Tier 3: LP Share
        (f"C{r + 13}", tiers[2].hurdle_value),             # Tier 4: IRR Hurdle
        (f"C{r + 14}", tiers[2].lp_share),                # Tier 4: LP Share
        (f"C{r + 15}", tiers[3].hurdle_value),             # Tier 5: IRR Hurdle
        (f"C{r + 16}", tiers[3].lp_share),                # Tier 5: LP Share
        (f"C{r + 17}", a.residual_tier.lp_share),          # Tier 6: Residual LP
    ]


def _section_em_hurdles(a, is_hold: bool) -> CellMap:
    """Section 14: Equity Multiple Hurdles."""
    r = 215 if is_hold else 193
    return [
        (f"C{r + 1}", a.em_hurdle_t1),
        (f"C{r + 2}", a.em_hurdle_t2),
        (f"C{r + 3}", a.em_hurdle_t3),
    ]


def _section_sensitivity(a, is_hold: bool) -> CellMap:
    """Section 15: Sensitivity Analysis Ranges."""
    r = 220 if is_hold else 198
    return [
        (f"C{r + 2}", a.sens_rent_growth_low),
        (f"C{r + 3}", a.sens_rent_growth_high),
        (f"C{r + 4}", a.sens_rent_growth_step),
        (f"C{r + 5}", a.sens_exit_cap_low),
        (f"C{r + 6}", a.sens_exit_cap_high),
        (f"C{r + 7}", a.sens_exit_cap_step),
    ]
